from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass
from decimal import Decimal
from typing import Iterable

from django.db import transaction
from django.utils.text import slugify

from .models import Categoria, Precio, Producto, ProductoVariante, Proveedor


HEADER_ALIASES = {
    "nombre": {"nombre", "name", "titulo", "producto", "nombre del producto"},
    "descripcion": {"descripcion", "description", "detalle", "descripción"},
    "sku": {"sku", "codigo", "codigo sku", "sku principal", "sku (obligatorio)", "sku obligatorio"},
    "stock": {"stock", "stock_actual", "inventario"},
    "stock_minimo": {"stock_minimo", "stock minimo", "minimo"},
    "peso": {"peso", "peso kg", "peso en kg", "peso (kg)", "peso en kg (kg)", "peso en kg (obligatorio)", "peso en kg.", "peso en kg (kg)"},
    "costo": {"costo", "costo_usd", "costo unitario"},
    "precio_minorista_ars": {"precio", "precio_minorista_ars", "precio ars", "precio venta", "precio minorista"},
    "precio_minorista_usd": {"precio_minorista_usd", "precio usd"},
    "precio_mayorista_ars": {"precio_mayorista_ars", "mayorista ars"},
    "precio_mayorista_usd": {"precio_mayorista_usd", "mayorista usd"},
    "precio_oferta_ars": {"oferta", "precio oferta", "precio promocional"},
    "categoria": {"categoria", "category", "categorias", "categorías > subcategorías > … > subcategorías", "categorías > subcategorías > … > subcategorías"},
    "proveedor": {"proveedor", "supplier", "proveedores", "vendor", "fabricante", "marca"},
    "codigo_barras": {"codigo barras", "barcode", "codigo_de_barras"},
    "atributo_1": {"opcion 1 valor", "opcion1", "atributo1", "color", "opción de variante #1", "opción variante 1", "opción de variante #1", "nombre de variante #1"},
    "atributo_2": {"opcion 2 valor", "opcion2", "atributo2", "talle", "capacidad", "opción de variante #2", "opción variante 2", "opción de variante #2", "nombre de variante #2"},
    "atributo_3": {"opcion 3 valor", "opcion3", "atributo3", "opción de variante #3", "opción variante 3", "nombre de variante #3"},
    "visibilidad": {"visibilidad", "visibilidad (visible o oculto)", "estado", "publicación", "visibilidad visible o oculto"},
}


@dataclass
class ImportResult:
    creados: int = 0
    actualizados: int = 0
    errores: list[str] = None
    omitidos: list[str] = None

    def __post_init__(self):
        if self.errores is None:
            self.errores = []
        if self.omitidos is None:
            self.omitidos = []


def _normalizar_header(header: str) -> str:
    normal = re.sub(r"\s+", " ", header.strip().lower())
    for canonical, aliases in HEADER_ALIASES.items():
        if normal in aliases:
            return canonical
    return normal


def _leer_csv(contenido: bytes) -> Iterable[dict[str, str]]:
    buffer = io.StringIO(contenido.decode("utf-8-sig"))
    reader = csv.DictReader(buffer)
    headers = reader.fieldnames or []
    mapping = {original: _normalizar_header(original) for original in headers}

    for row in reader:
        data: dict[str, str] = {}
        for original, value in row.items():
            if not original:
                continue
            canonical = mapping.get(original)
            if not canonical:
                continue
            if isinstance(value, str):
                value = value.strip()
            data[canonical] = value
        yield data


def _leer_xlsx(contenido: bytes) -> Iterable[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependencia opcional
        raise RuntimeError("openpyxl es requerido para importar archivos XLSX") from exc

    buffer = io.BytesIO(contenido)
    wb = load_workbook(buffer, read_only=True)
    sheet = wb.active
    rows = sheet.iter_rows(values_only=True)
    headers = [str(h or "").strip() for h in next(rows)]
    headers_normalizados = [_normalizar_header(h) for h in headers]

    for row in rows:
        data = {}
        for header, value in zip(headers_normalizados, row):
            if header:
                data[header] = str(value).strip() if value is not None else ""
        yield data


def _parse_decimal(value: str) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, (int, float, Decimal)):
        try:
            return Decimal(str(value))
        except Exception:
            return None
    value = str(value).strip()
    if not value:
        return None
    clean = value.replace("$", "").replace(" ", "").replace("\u00a0", "")
    # Si contiene tanto punto como coma, asumir formato latino (1.234,56)
    if "." in clean and "," in clean:
        clean = clean.replace(".", "").replace(",", ".")
    elif clean.count(",") > 1:
        # Demasiadas comas -> probablemente separadores de miles en latino
        clean = clean.replace(",", "")
    elif "," in clean:
        clean = clean.replace(",", ".")
    else:
        # Solo puntos: si hay más de uno, remover todos menos el último
        if clean.count(".") > 1:
            parts = clean.split(".")
            decimal_part = parts[-1]
            entero = "".join(parts[:-1])
            clean = f"{entero}.{decimal_part}"
    try:
        return Decimal(clean)
    except Exception:
        return None


def _obtener_categoria(nombre: str | None, parent: Categoria | None = None) -> Categoria | None:
    """
    Obtiene o crea una categoría. Soporta jerarquía usando el formato:
    "Categoría > Subcategoría > Sub-subcategoría"
    """
    if not nombre:
        return None
    
    nombre = nombre.strip()
    
    # Si el nombre contiene ">", es una jerarquía
    if ">" in nombre:
        partes = [p.strip() for p in nombre.split(">")]
        categoria_actual = None
        
        for parte in partes:
            if not parte:
                continue
            
            # Buscar o crear la categoría con el parent correspondiente
            categoria_actual, _ = Categoria.objects.get_or_create(
                nombre=parte,
                parent=categoria_actual,
                defaults={
                    "descripcion": "",
                }
            )
        
        return categoria_actual
    else:
        # Categoría simple (sin jerarquía)
        categoria, _ = Categoria.objects.get_or_create(
            nombre=nombre,
            parent=parent,
            defaults={
                "descripcion": "",
            }
        )
        return categoria


def _obtener_proveedor(nombre: str | None) -> Proveedor | None:
    if not nombre:
        return None
    return Proveedor.objects.get_or_create(nombre=nombre.strip())[0]


@transaction.atomic
def importar_catalogo_desde_archivo(archivo, actualizar: bool = True) -> ImportResult:
    nombre = archivo.name.lower()
    contenido = archivo.read()

    if nombre.endswith(".csv"):
        filas = list(_leer_csv(contenido))
    elif nombre.endswith((".xlsx", ".xls")):
        filas = list(_leer_xlsx(contenido))
    else:
        raise ValueError("Formato no soportado. Usá CSV o XLSX.")

    resultado = ImportResult()
    
    from core.db_inspector import column_exists
    from django.db import connection
    
    tiene_col_stock = column_exists("inventario_productovariante", "stock")
    tiene_col_peso = column_exists("inventario_productovariante", "peso")

    for idx, fila in enumerate(filas, start=2):  # encabezado es la fila 1
        data = {k: fila.get(k, "") for k in HEADER_ALIASES}
        # Para MercadoLibre: SKU es obligatorio, si no está, generamos desde nombre + variantes
        sku = (data.get("sku") or "").strip()
        
        # Si hay SKU en la celda, verificar que no esté duplicado
        if sku:
            # Truncar SKU si es demasiado largo (max 64 caracteres)
            if len(sku) > 64:
                sku_original = sku
                sku = sku[:64]
                resultado.errores.append(f"Fila {idx}: SKU '{sku_original}' truncado a '{sku}' (máximo 64 caracteres).")
            
            # Verificar si el SKU ya existe
            if ProductoVariante.objects.filter(sku=sku).exists():
                # Si existe y no queremos actualizar, generar uno nuevo
                if not actualizar:
                    # Generar SKU único agregando sufijo numérico
                    base_sku = sku
                    # Asegurar que el base_sku deje espacio para el contador
                    if len(base_sku) > 60:
                        base_sku = base_sku[:60]
                    contador = 1
                    while ProductoVariante.objects.filter(sku=sku).exists():
                        sku_con_contador = f"{base_sku}-{contador}"
                        if len(sku_con_contador) > 64:
                            # Si es muy largo, truncar más el base
                            max_base_len = 64 - len(f"-{contador}")
                            base_sku = base_sku[:max_base_len]
                            sku_con_contador = f"{base_sku}-{contador}"
                        sku = sku_con_contador
                        contador += 1
                    resultado.errores.append(f"Fila {idx}: SKU '{base_sku}' ya existe, usando '{sku}' en su lugar.")
        else:
            # Auto-generar SKU desde nombre + variantes
            nombre = data.get("nombre", "")
            attr1 = data.get("atributo_1", "")
            attr2 = data.get("atributo_2", "")
            attr3 = data.get("atributo_3", "")
            sku_base = slugify(f"{nombre}-{attr1}-{attr2}-{attr3}".strip("-")).upper()
            
            if not sku_base:
                resultado.errores.append(f"Fila {idx}: sin SKU ni nombre.")
                continue
            
            # Truncar SKU base a 60 caracteres para dejar espacio al contador
            if len(sku_base) > 60:
                sku_base = sku_base[:60]
            
            # Verificar que el SKU generado no esté duplicado
            sku = sku_base
            contador = 1
            while ProductoVariante.objects.filter(sku=sku).exists():
                # Asegurar que el SKU con contador no exceda 64 caracteres
                sku_con_contador = f"{sku_base}-{contador}"
                if len(sku_con_contador) > 64:
                    # Si es muy largo, truncar más el base
                    max_base_len = 64 - len(f"-{contador}")
                    sku_base = sku_base[:max_base_len]
                    sku_con_contador = f"{sku_base}-{contador}"
                sku = sku_con_contador
                contador += 1
            
            # Asegurar que el SKU final no exceda 64 caracteres
            if len(sku) > 64:
                sku = sku[:64]

        # Validaciones básicas de stock
        try:
            stock_val = int(data.get("stock") or 0)
            stock_min_val = int(data.get("stock_minimo") or 0)
            if stock_val < 0 or stock_min_val < 0:
                resultado.errores.append(f"Fila {idx}: stock o stock mínimo negativo.")
                continue
        except Exception:
            resultado.errores.append(f"Fila {idx}: stock inválido.")
            continue

        peso_valor = _parse_decimal(data.get("peso"))
        if peso_valor is None:
            peso_valor = Decimal("0")

        visibilidad = (data.get("visibilidad") or "").strip().lower()
        is_visible = True
        if visibilidad:
            if any(term in visibilidad for term in ["ocult", "hidden", "no", "desactiv", "draft", "inactivo", "0"]):
                is_visible = False

        categoria_obj = _obtener_categoria(data.get("categoria"))
        
        # Filtrar iPhones (categoría "Celulares") - no importar
        if categoria_obj and categoria_obj.nombre.lower() == "celulares":
            resultado.omitidos.append(f"Fila {idx}: Producto '{data.get('nombre', sku)}' es un iPhone (categoría Celulares) - omitido")
            continue
        
        producto_defaults = {
            "descripcion": data.get("descripcion", ""),
            "categoria": categoria_obj,
            "proveedor": _obtener_proveedor(data.get("proveedor")),
            "codigo_barras": data.get("codigo_barras", ""),
            "activo": is_visible,
            "estado": "ACTIVO",
        }

        # Usar filter().first() para evitar el error "get() returned more than one Producto"
        nombre_producto = data.get("nombre") or sku
        producto = Producto.objects.filter(nombre=nombre_producto).first()
        
        if producto:
            # Producto existe, actualizar si es necesario
            if actualizar:
                for campo, valor in producto_defaults.items():
                    if valor and getattr(producto, campo) != valor:
                        setattr(producto, campo, valor)
                producto.save()
        else:
            # Crear nuevo producto
            producto = Producto.objects.create(nombre=nombre_producto, **producto_defaults)

        defaults_variante = {
            "atributo_1": data.get("atributo_1", ""),
            "atributo_2": data.get("atributo_2", ""),
            "stock_actual": stock_val,
            "stock_minimo": stock_min_val,
            "peso": peso_valor,
            "activo": is_visible,
        }
        # Si hay atributo_3, lo agregamos a atributo_2 o creamos un campo combinado
        if data.get("atributo_3"):
            if defaults_variante["atributo_2"]:
                defaults_variante["atributo_2"] = f"{defaults_variante['atributo_2']} / {data.get('atributo_3')}"
            else:
                defaults_variante["atributo_2"] = data.get("atributo_3", "")
        
        # Generar nombre_variante único basado en atributos o SKU
        attr1 = defaults_variante.get("atributo_1", "").strip()
        attr2 = defaults_variante.get("atributo_2", "").strip()
        
        if attr1 or attr2:
            # Si hay atributos, combinarlos
            partes = [p for p in [attr1, attr2] if p]
            nombre_variante_base = " / ".join(partes)
        else:
            # Si no hay atributos, usar el SKU o dejar vacío
            nombre_variante_base = ""
        
        # Asegurar que el nombre_variante sea único para este producto
        # Primero verificar si ya existe una variante con este SKU
        variante_existente_por_sku = ProductoVariante.objects.filter(sku=sku).first()
        
        nombre_variante = nombre_variante_base
        contador_nombre = 1
        # Verificar que no exista otra variante del mismo producto con el mismo nombre_variante
        # (excluyendo la variante actual si existe)
        while True:
            query = ProductoVariante.objects.filter(producto=producto, nombre_variante=nombre_variante)
            if variante_existente_por_sku:
                query = query.exclude(pk=variante_existente_por_sku.pk)
            if not query.exists():
                break
            nombre_variante = f"{nombre_variante_base} ({contador_nombre})"
            contador_nombre += 1

        # Verificar si existen columnas legacy para compatibilidad
        if tiene_col_stock:
            # Si existe el campo stock, usar SQL directo para insertar
            from django.utils import timezone
            ahora = timezone.now()
            
            # Usar la variante que ya verificamos antes
            variante_existente = variante_existente_por_sku
            
            if variante_existente:
                if actualizar:
                    for campo, valor in defaults_variante.items():
                        setattr(variante_existente, campo, valor)
                    # Actualizar también el nombre_variante
                    variante_existente.nombre_variante = nombre_variante
                    if variante_existente.producto_id != producto.id:
                        resultado.errores.append(f"Fila {idx}: el SKU {sku} ya está asociado a otro producto.")
                        continue
                    variante_existente.save()
                    # Actualizar también el campo stock antiguo
                    with connection.cursor() as cursor:
                        if tiene_col_peso:
                            cursor.execute(
                                "UPDATE inventario_productovariante SET stock = %s, peso = %s, nombre_variante = %s WHERE id = %s",
                                [stock_val, peso_valor, nombre_variante, variante_existente.pk]
                            )
                        else:
                            cursor.execute(
                                "UPDATE inventario_productovariante SET stock = %s, nombre_variante = %s WHERE id = %s",
                                [stock_val, nombre_variante, variante_existente.pk]
                            )
                    resultado.actualizados += 1
                else:
                    resultado.errores.append(f"Fila {idx}: el SKU {sku} ya existe y no se actualizó.")
                    continue
                variante = variante_existente
            else:
                # Crear nueva variante con SQL directo
                with connection.cursor() as cursor:
                    columnas = [
                        "producto_id", "sku", "nombre_variante", "codigo_barras", "qr_code",
                        "atributo_1", "atributo_2", "stock_actual", "stock_minimo", "stock"
                    ]
                    valores = [
                        producto.pk, sku, nombre_variante,
                        data.get("codigo_barras") or None,
                        None,  # qr_code
                        defaults_variante.get("atributo_1", ""),
                        defaults_variante.get("atributo_2", ""),
                        stock_val, stock_min_val, stock_val
                    ]
                    if tiene_col_peso:
                        columnas.append("peso")
                        valores.append(peso_valor)
                    
                    # Verificar si existe el campo costo
                    tiene_costo = column_exists("inventario_productovariante", "costo")
                    if tiene_costo:
                        columnas.append("costo")
                        valores.append(0)  # Valor por defecto para costo
                    
                    columnas.extend(["activo", "creado", "actualizado"])
                    valores.extend([1 if is_visible else 0, ahora, ahora])
                    placeholders = ", ".join(["%s"] * len(valores))
                    cursor.execute(
                        f"""
                        INSERT INTO inventario_productovariante 
                        ({", ".join(columnas)})
                        VALUES ({placeholders})
                        """,
                        valores
                    )
                    variante_id = cursor.lastrowid
                variante = ProductoVariante.objects.get(pk=variante_id)
                resultado.creados += 1
        else:
            # Si no existe el campo stock, usar el método normal
            defaults_variante["nombre_variante"] = nombre_variante
            variante, creada = ProductoVariante.objects.get_or_create(sku=sku, defaults={"producto": producto, **defaults_variante})
            if creada:
                resultado.creados += 1
            else:
                if actualizar:
                    for campo, valor in defaults_variante.items():
                        setattr(variante, campo, valor)
                    variante.nombre_variante = nombre_variante
                    if variante.producto_id != producto.id:
                        resultado.errores.append(f"Fila {idx}: el SKU {sku} ya está asociado a otro producto.")
                        continue
                    variante.save()
                    resultado.actualizados += 1
                else:
                    resultado.errores.append(f"Fila {idx}: el SKU {sku} ya existe y no se actualizó.")
                    continue

        precio_minorista = _parse_decimal(data.get("precio_minorista_ars"))
        if precio_minorista is None:
            precio_minorista = _parse_decimal(data.get("precio_oferta_ars"))

        precios = {
            (Precio.Tipo.MINORISTA, Precio.Moneda.ARS): precio_minorista,
            (Precio.Tipo.MINORISTA, Precio.Moneda.USD): _parse_decimal(data.get("precio_minorista_usd")),
            (Precio.Tipo.MAYORISTA, Precio.Moneda.ARS): _parse_decimal(data.get("precio_mayorista_ars")),
            (Precio.Tipo.MAYORISTA, Precio.Moneda.USD): _parse_decimal(data.get("precio_mayorista_usd")),
        }

        # Verificar qué campos existen en la tabla inventario_precio
        tiene_tipo_precio = column_exists("inventario_precio", "tipo_precio")
        tiene_costo = column_exists("inventario_precio", "costo")
        tiene_precio_venta_normal = column_exists("inventario_precio", "precio_venta_normal")
        tiene_precio_venta_minimo = column_exists("inventario_precio", "precio_venta_minimo")
        tiene_precio_venta_descuento = column_exists("inventario_precio", "precio_venta_descuento")
        
        # Obtener costo si está disponible
        costo_valor = _parse_decimal(data.get("costo")) or Decimal("0")

        for (tipo, moneda), valor in precios.items():
            if valor is None:
                continue
            
            # Verificar si existe el precio
            precio_existente = Precio.objects.filter(
                variante=variante,
                tipo=tipo,
                moneda=moneda
            ).first()
            
            if precio_existente:
                # Solo actualizar si actualizar=True
                if actualizar:
                    # Actualizar el precio existente
                    update_fields = ["precio = %s", "activo = %s", "actualizado = NOW()"]
                    update_values = [valor, True]
                    
                    if tiene_tipo_precio:
                        update_fields.append("tipo_precio = %s")
                        update_values.append(tipo)
                    
                    if tiene_costo:
                        update_fields.append("costo = %s")
                        update_values.append(costo_valor)
                    
                    if tiene_precio_venta_normal:
                        update_fields.append("precio_venta_normal = %s")
                        update_values.append(valor)
                    
                    if tiene_precio_venta_minimo:
                        update_fields.append("precio_venta_minimo = %s")
                        update_values.append(valor)
                    
                    update_values.append(precio_existente.pk)
                    
                    with connection.cursor() as cursor:
                        cursor.execute(f"""
                            UPDATE inventario_precio 
                            SET {', '.join(update_fields)}
                            WHERE id = %s
                        """, update_values)
            else:
                # Crear nuevo precio con SQL directo
                insert_fields = ["variante_id", "tipo", "moneda", "precio", "activo"]
                insert_values = [variante.pk, tipo, moneda, valor, True]
                
                if tiene_tipo_precio:
                    insert_fields.append("tipo_precio")
                    insert_values.append(tipo)
                
                if tiene_costo:
                    insert_fields.append("costo")
                    insert_values.append(costo_valor)
                
                if tiene_precio_venta_normal:
                    insert_fields.append("precio_venta_normal")
                    insert_values.append(valor)
                
                if tiene_precio_venta_minimo:
                    insert_fields.append("precio_venta_minimo")
                    insert_values.append(valor)
                
                if tiene_precio_venta_descuento:
                    insert_fields.append("precio_venta_descuento")
                    insert_values.append(None)
                
                # Agregar timestamps al final (usando NOW() directamente en SQL)
                insert_fields.extend(["creado", "actualizado"])
                
                with connection.cursor() as cursor:
                    placeholders = ", ".join(["%s"] * len(insert_values) + ["NOW()", "NOW()"])
                    cursor.execute(f"""
                        INSERT INTO inventario_precio 
                        ({', '.join(insert_fields)})
                        VALUES ({placeholders})
                    """, insert_values)

    return resultado


def exportar_catalogo_a_excel(categoria_id=None) -> io.BytesIO:
    """Exporta el catálogo en formato Excel compatible con MercadoLibre."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        raise RuntimeError("openpyxl es requerido para exportar a Excel")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Catálogo"
    
    # Headers estilo MercadoLibre
    headers = [
        "SKU (OBLIGATORIO)",
        "Nombre del producto",
        "Precio Minorista ARS",
        "Precio Mayorista ARS",
        "Costo",
        "Oferta",
        "Stock",
        "Visibilidad (Visible o Oculto)",
        "Descripción",
        "Peso en KG",
        "Alto en CM",
        "Ancho en CM",
        "Profundidad en CM",
        "Nombre de variante #1",
        "Opción de variante #1",
        "Nombre de variante #2",
        "Opción de variante #2",
        "Nombre de variante #3",
        "Opción de variante #3",
        "Categorías > Subcategorías > … > Subcategorías",
    ]
    
    # Estilo de headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Filtrar iPhones (categoría "Celulares") - solo exportar otros productos
    variantes = ProductoVariante.objects.select_related("producto", "producto__proveedor").prefetch_related("precios").exclude(
        producto__categoria__nombre__iexact="celulares"
    )
    
    # Si se especifica una categoría, filtrar por ella y sus subcategorías
    if categoria_id:
        from .models import Categoria
        try:
            categoria = Categoria.objects.get(pk=categoria_id)
            # Obtener todas las subcategorías recursivamente
            def obtener_subcategorias(cat):
                subcats = [cat]
                for subcat in cat.subcategorias.all():
                    subcats.extend(obtener_subcategorias(subcat))
                return subcats
            
            categorias_ids = [c.id for c in obtener_subcategorias(categoria)]
            variantes = variantes.filter(producto__categoria_id__in=categorias_ids)
        except Categoria.DoesNotExist:
            pass
    
    row = 2
    for variante in variantes:
        def precio(tipo, moneda):
            registro = variante.precios.filter(tipo=tipo, moneda=moneda, activo=True).order_by("-actualizado").first()
            return registro.precio if registro else ""
        
        precio_minorista_ars = precio(Precio.Tipo.MINORISTA, Precio.Moneda.ARS)
        precio_mayorista_ars = precio(Precio.Tipo.MAYORISTA, Precio.Moneda.ARS)
        precio_oferta = precio(Precio.Tipo.MINORISTA, Precio.Moneda.ARS)  # Puedes ajustar esto
        
        # Obtener costo desde la variante o desde precio
        costo_valor = ""
        if hasattr(variante, 'costo') and variante.costo:
            costo_valor = variante.costo
        else:
            # Intentar obtener desde precio
            precio_obj = variante.precios.filter(activo=True).first()
            if precio_obj and hasattr(precio_obj, 'costo') and precio_obj.costo:
                costo_valor = precio_obj.costo
        
        visibilidad = "Visible" if variante.activo and variante.producto.activo else "Oculto"
        categoria = variante.producto.categoria.nombre_completo if variante.producto.categoria else ""
        
        ws.cell(row=row, column=1, value=variante.sku)
        ws.cell(row=row, column=2, value=variante.producto.nombre)
        ws.cell(row=row, column=3, value=precio_minorista_ars)
        ws.cell(row=row, column=4, value=precio_mayorista_ars)
        ws.cell(row=row, column=5, value=costo_valor)
        ws.cell(row=row, column=6, value=precio_oferta)
        ws.cell(row=row, column=7, value=variante.stock_actual)
        ws.cell(row=row, column=8, value=visibilidad)
        ws.cell(row=row, column=9, value=variante.producto.descripcion or "")
        ws.cell(row=row, column=14, value=variante.atributo_1 or "")
        ws.cell(row=row, column=15, value=variante.atributo_1 or "")
        ws.cell(row=row, column=16, value=variante.atributo_2 or "")
        ws.cell(row=row, column=17, value=variante.atributo_2 or "")
        ws.cell(row=row, column=20, value=categoria)
        row += 1
    
    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def exportar_catalogo_a_csv(categoria_id=None) -> io.BytesIO:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(
        [
            "SKU (OBLIGATORIO)",
            "Nombre del producto",
            "Precio Minorista ARS",
            "Precio Mayorista ARS",
            "Costo",
            "Oferta",
            "Stock",
            "Visibilidad (Visible o Oculto)",
            "Descripción",
            "Peso en KG",
            "Alto en CM",
            "Ancho en CM",
            "Profundidad en CM",
            "Nombre de variante #1",
            "Opción de variante #1",
            "Nombre de variante #2",
            "Opción de variante #2",
            "Nombre de variante #3",
            "Opción de variante #3",
            "Categorías > Subcategorías > … > Subcategorías",
        ]
    )

    # Filtrar iPhones (categoría "Celulares") - solo exportar otros productos
    variantes = ProductoVariante.objects.select_related("producto", "producto__proveedor").prefetch_related("precios").exclude(
        producto__categoria__nombre__iexact="celulares"
    )
    
    # Si se especifica una categoría, filtrar por ella y sus subcategorías
    if categoria_id:
        from .models import Categoria
        try:
            categoria = Categoria.objects.get(pk=categoria_id)
            # Obtener todas las subcategorías recursivamente
            def obtener_subcategorias(cat):
                subcats = [cat]
                for subcat in cat.subcategorias.all():
                    subcats.extend(obtener_subcategorias(subcat))
                return subcats
            
            categorias_ids = [c.id for c in obtener_subcategorias(categoria)]
            variantes = variantes.filter(producto__categoria_id__in=categorias_ids)
        except Categoria.DoesNotExist:
            pass
    
    for variante in variantes:
        def precio(tipo, moneda):
            registro = variante.precios.filter(tipo=tipo, moneda=moneda, activo=True).order_by("-actualizado").first()
            return registro.precio if registro else ""

        precio_minorista_ars = precio(Precio.Tipo.MINORISTA, Precio.Moneda.ARS)
        precio_mayorista_ars = precio(Precio.Tipo.MAYORISTA, Precio.Moneda.ARS)
        precio_oferta = precio(Precio.Tipo.MINORISTA, Precio.Moneda.ARS)  # Ajustar según necesidad
        
        # Obtener costo desde la variante o desde precio
        costo_valor = ""
        if hasattr(variante, 'costo') and variante.costo:
            costo_valor = variante.costo
        else:
            # Intentar obtener desde precio
            precio_obj = variante.precios.filter(activo=True).first()
            if precio_obj and hasattr(precio_obj, 'costo') and precio_obj.costo:
                costo_valor = precio_obj.costo
        
        visibilidad = "Visible" if variante.activo and variante.producto.activo else "Oculto"
        categoria = variante.producto.categoria.nombre_completo if variante.producto.categoria else ""
        
        writer.writerow(
            [
                variante.sku,
                variante.producto.nombre,
                precio_minorista_ars,
                precio_mayorista_ars,
                costo_valor,
                precio_oferta,
                variante.stock_actual,
                visibilidad,
                variante.producto.descripcion or "",
                "",  # Peso en KG
                "",  # Alto en CM
                "",  # Ancho en CM
                "",  # Profundidad en CM
                variante.atributo_1 or "",
                variante.atributo_1 or "",
                variante.atributo_2 or "",
                variante.atributo_2 or "",
                "",  # Variante #3
                "",  # Opción variante #3
                categoria,
            ]
        )

    byte_buffer = io.BytesIO()
    byte_buffer.write(buffer.getvalue().encode("utf-8"))
    byte_buffer.seek(0)
    return byte_buffer


def generar_excel_vacio() -> io.BytesIO:
    """Genera un Excel vacío/template para cargar productos nuevos."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        raise RuntimeError("openpyxl es requerido para exportar a Excel")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Catálogo"
    
    # Headers estilo MercadoLibre
    headers = [
        "SKU (OBLIGATORIO)",
        "Nombre del producto",
        "Precio Minorista ARS",
        "Precio Mayorista ARS",
        "Costo",
        "Oferta",
        "Stock",
        "Visibilidad (Visible o Oculto)",
        "Descripción",
        "Peso en KG",
        "Alto en CM",
        "Ancho en CM",
        "Profundidad en CM",
        "Nombre de variante #1",
        "Opción de variante #1",
        "Nombre de variante #2",
        "Opción de variante #2",
        "Nombre de variante #3",
        "Opción de variante #3",
        "Categorías > Subcategorías > … > Subcategorías",
    ]
    
    # Estilo de headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Agregar una fila de ejemplo (opcional, puede comentarse)
    # ws.cell(row=2, column=1, value="EJEMPLO-SKU-001")
    # ws.cell(row=2, column=2, value="Producto de ejemplo")
    # ws.cell(row=2, column=3, value="10000")
    # ws.cell(row=2, column=4, value="8000")
    # ws.cell(row=2, column=5, value="5000")
    # ws.cell(row=2, column=7, value="10")
    # ws.cell(row=2, column=8, value="Visible")
    
    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
